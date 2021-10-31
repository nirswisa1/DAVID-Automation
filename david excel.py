from openpyxl import load_workbook
import pandas as pd
import matplotlib.pyplot as plt
from tkinter import *

# GUI
root = Tk()
root.title('Nir.S - DAVID Excel')
fname = ''


def myClick():
    global fname
    fname = fnameE.get()
    root.destroy()
    return fname


# slots
fnameE = Entry(root, width=50, borderwidth=2)
fnameE.grid(row=0, column=1)

# labels
label1 = Label(root, text='File name: ')
label1.grid(row=0, column=0)

# Button
b = Button(root, text="Run", command=myClick)
b.grid(row=1, column=1)

root.mainloop()

# excel
wb = load_workbook(fname + '.xlsx')
ws = wb.active
added = []


def funcx(current):
    global added
    global fname
    gene_names = ws["F" + str(current)].value  # ['QIHEL'...]
    for row in range(current, 50):
        cur_check = ws["F" + str(row)].value  # ['QIFFF'...]
        cur_gene_name = ws["B" + str(row)].value  # ['GO:302198...']
        if cur_check is None:
            break
        if cur_check == gene_names and cur_gene_name not in added:
            added.append(cur_gene_name)
            if ws["O" + str(current)].value is None:
                if cur_gene_name[0] == "G":
                    ws["O" + str(current)].value = (str(ws["O" + str(current)].value) + ", " + cur_gene_name[
                                                                                               cur_gene_name.find(
                                                                                                   '~') + 1:])[5:]
                elif cur_gene_name[0] == "m":
                    ws["O" + str(current)].value = (str(ws["O" + str(current)].value) + ", " + cur_gene_name[
                                                                                               cur_gene_name.find(
                                                                                                   ':') + 1:])[5:]
            elif cur_gene_name[0] == "G":
                ws["O" + str(current)].value = str(ws["O" + str(current)].value) + ", " + cur_gene_name[
                                                                                          cur_gene_name.find('~') + 1:]
            elif cur_gene_name[0] == "m":
                ws["O" + str(current)].value = str(ws["O" + str(current)].value) + ", " + cur_gene_name[
                                                                                          cur_gene_name.find(':') + 1:]

    wb.save(fname + '.xlsx')

#Main loop
Lcolors = []
current = 1
colorNum = -1
for cell in ws['F']:
    if cell.value is None:
        added = []
        current += 1
    elif cell.value == "Genes":
        added = []
        current += 1
        colorNum += 1
    else:
        Lcolors.append(colorNum)
        funcx(current)
        current += 1
#Plot
names = []
numbers = []
for num, cell in enumerate(ws["O"]):
    if cell.value is not None:
        names.append(str(cell.value))
        numbers.append(ws["C" + str(num + 1)].value)

colors = ['teal', 'powderblue', 'lightgreen', 'cadetblue', 'paleturquoise', 'palegreen']
new_data = {'names': list(reversed(names)), 'numbers': list(reversed(numbers))}
df = pd.DataFrame.from_dict(new_data);
df
plt.barh(df.names, df.numbers, color=[colors[i] for i in Lcolors])
plt.show()
